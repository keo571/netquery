#!/usr/bin/env python3
"""
Add a 'suggested_queries' sheet to sample_schema.xlsx with sample questions.
"""
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

def add_suggested_queries_sheet(excel_path: str):
    """Add suggested_queries sheet to Excel file."""

    # Define suggested queries for the sample database
    # Today is November 28, 2025 - time series queries should use last 30 days
    suggested_queries = [
        # Bar chart queries (count/aggregation by category)
        "Show count of load balancers by datacenter",
        "Show count of load balancers by status",
        "Show count of backend servers by health status",
        "Show count of virtual IPs by protocol",

        # Pie chart queries (distribution/proportions)
        "Show distribution of load balancers by status",
        "Show distribution of backend servers by health status",
        "Show distribution of virtual IPs by protocol",

        # Line chart queries (time series - last 30 days: Oct 29 - Nov 28, 2025)
        "Show load balancer performance stats over the last 30 days",
        "Show requests per second trend for the last 30 days",
        "Show connection count trend over the last 30 days",
        "Show bandwidth usage over the last 30 days",

        # Table/list queries (for reference)
        "Show all active load balancers",
        "List all virtual IPs with their backend servers",
        "Show backend servers that are currently unhealthy",
        "List load balancers with their total number of backend servers"
    ]

    # Create DataFrame
    df = pd.DataFrame({'query': suggested_queries})

    # Load existing workbook
    excel_path = Path(excel_path)
    book = load_workbook(excel_path)

    # Check if suggested_queries sheet already exists
    if 'suggested_queries' in book.sheetnames:
        print(f"Sheet 'suggested_queries' already exists in {excel_path}. Removing it...")
        del book['suggested_queries']

    # Create the new sheet
    ws = book.create_sheet('suggested_queries')

    # Write header
    ws.cell(row=1, column=1, value='query')

    # Write data
    for idx, query in enumerate(suggested_queries, start=2):
        ws.cell(row=idx, column=1, value=query)

    # Save workbook
    book.save(excel_path)

    print(f"✅ Added 'suggested_queries' sheet to {excel_path}")
    print(f"   Added {len(suggested_queries)} suggested queries")
    print("\nSample queries:")
    for i, query in enumerate(suggested_queries[:5], 1):
        print(f"   {i}. {query}")
    print(f"   ... and {len(suggested_queries) - 5} more")

if __name__ == "__main__":
    schema_file = "schema_files/sample_schema.xlsx"
    add_suggested_queries_sheet(schema_file)
